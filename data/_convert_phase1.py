# -*- coding: utf-8 -*-
"""
阶段一 xlsx → JSON 转换器
输入：/tmp/阶段一_view/阶段一/数据爬取&清理/*.xlsx
输出：当前目录 phase1/ 下的 JSON 文件，供前端 D3 viz 使用。
"""
import json
import os
from pathlib import Path
import openpyxl

SRC = Path("/tmp/阶段一_view/阶段一/数据爬取&清理")
OUT = Path(__file__).parent / "phase1"
OUT.mkdir(parents=True, exist_ok=True)


def load(name):
    return openpyxl.load_workbook(SRC / name, data_only=True)


def rows(ws, header_row=1, start=None):
    """Yield dict rows. header_row 是表头行号（1-indexed），start 是数据开始行号。"""
    headers = [c.value for c in ws[header_row]]
    start = start or header_row + 1
    for r in ws.iter_rows(min_row=start, values_only=True):
        if all(v is None for v in r):
            continue
        yield {h: v for h, v in zip(headers, r) if h is not None}


def dump(name, data):
    path = OUT / name
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)
    print(f"  ✓ {name}: {path.stat().st_size/1024:.1f} KB")


# ------------------------------------------------------------
# 1. 商场基本信息.xlsx → malls.json
# 这份 xlsx 表头在第二行（第一行是合并的标题 'Shanghai Malls'）
# ------------------------------------------------------------
print("\n[1/7] 商场基本信息 → malls.json")
wb = load("商场基本信息.xlsx")
ws = wb["shanghai-malls"]
malls = []
for r in rows(ws, header_row=2, start=3):
    malls.append({
        "id": r.get("序号"),
        "name": r.get("标准名称"),
        "poi": r.get("POI_IDs"),
        "address": r.get("地址"),
        "district": r.get("区"),
        "lng": r.get("经度"),
        "lat": r.get("纬度"),
        "distance_km": r.get("距人民广场km"),
        "circle": r.get("空间圈层(近似)"),
        "positioning": r.get("档次定位"),
    })
print(f"  共 {len(malls)} 个商场")
dump("malls.json", malls)


# ------------------------------------------------------------
# 2. 商场品牌同质化.xlsx → mhi_brand.json + similarity_brand.json
# ------------------------------------------------------------
print("\n[2/7] 商场品牌同质化 → mhi_brand.json + similarity_brand.json")
wb = load("商场品牌同质化.xlsx")
ws_sum = wb["MHI汇总"]
mhi_brand = []
for r in rows(ws_sum):
    mhi_brand.append({
        "rank": r.get("MHI排名"),
        "id": r.get("mall_order"),
        "name": r.get("mall_name"),
        "brand_count": r.get("brand_count"),
        "MHI_i": r.get("MHI_i"),
        "avg_shared": r.get("平均共有品牌数"),
        "avg_union": r.get("平均并集品牌数"),
        "top_sim": r.get("最高两两相似度"),
        "top_sim_mall": r.get("最相似商场"),
        "bottom_sim": r.get("最低两两相似度"),
        "bottom_sim_mall": r.get("最不相似商场"),
    })
print(f"  MHI_i 汇总 {len(mhi_brand)} 行")
dump("mhi_brand.json", mhi_brand)

# 80×80 品牌相似度矩阵
ws_mat = wb["相似度矩阵"]
header = [c.value for c in ws_mat[1]]
mall_cols = header[3:]
sim_brand_matrix = []
mall_names_brand = []
for r in ws_mat.iter_rows(min_row=2, values_only=True):
    if r[0] is None:
        continue
    name = r[1]
    mall_names_brand.append(name)
    sim_brand_matrix.append([
        (v if v is not None else None) for v in r[3:]
    ])
dump("similarity_brand.json", {
    "malls": mall_names_brand,
    "matrix": sim_brand_matrix,
})


# ------------------------------------------------------------
# 3. 商场功能同质化.xlsx → mhi_function.json + similarity_function.json + function_vectors.json
# ------------------------------------------------------------
print("\n[3/7] 商场功能同质化 → mhi_function.json + similarity_function.json + function_vectors.json")
wb = load("商场功能同质化.xlsx")

# MHI_f 汇总
ws_sum = wb["MHI_f汇总"]
mhi_func = []
for r in rows(ws_sum):
    mhi_func.append({
        "rank": r.get("MHI_f排名"),
        "id": r.get("mall_order"),
        "name": r.get("mall_name"),
        "total_brand_count": r.get("总功能品牌数"),
        "function_count": r.get("覆盖功能数"),
        "vector_norm": r.get("向量模长"),
        "MHI_f": r.get("MHI_f"),
        "top_sim": r.get("最高两两功能相似度"),
        "top_sim_mall": r.get("最相似商场"),
        "bottom_sim": r.get("最低两两功能相似度"),
        "bottom_sim_mall": r.get("最不相似商场"),
    })
dump("mhi_function.json", mhi_func)

# 80×80 功能相似度矩阵
ws_mat = wb["功能相似度矩阵"]
sim_func_matrix = []
mall_names_func = []
for r in ws_mat.iter_rows(min_row=2, values_only=True):
    if r[0] is None:
        continue
    mall_names_func.append(r[1])
    sim_func_matrix.append(list(r[3:]))
dump("similarity_function.json", {
    "malls": mall_names_func,
    "matrix": sim_func_matrix,
})

# 80 × 14 功能向量
ws_vec = wb["功能向量矩阵"]
header = [c.value for c in ws_vec[1]]
function_axes = header[5:]  # 14 个业态分类
vectors = []
for r in ws_vec.iter_rows(min_row=2, values_only=True):
    if r[0] is None:
        continue
    vectors.append({
        "id": r[0],
        "name": r[1],
        "total_brand_count": r[2],
        "function_count": r[3],
        "vector_norm": r[4],
        "vector": list(r[5:]),
    })
dump("function_vectors.json", {
    "axes": function_axes,
    "data": vectors,
})


# ------------------------------------------------------------
# 4. 商场复合MHI指数.xlsx → mhi_composite.json
# ------------------------------------------------------------
print("\n[4/7] 商场复合MHI指数 → mhi_composite.json")
wb = load("商场复合MHI指数.xlsx")
ws = wb["复合MHI指数"]
composite = []
for r in rows(ws):
    composite.append({
        "rank": r.get("H_rank"),
        "id": r.get("mall_order"),
        "name": r.get("mall_name"),
        "district": r.get("district"),
        "circle": r.get("circle"),
        "positioning": r.get("positioning"),
        "brand_count": r.get("brand_count"),
        "total_function_brand_count": r.get("总功能品牌数"),
        "function_count": r.get("覆盖功能数"),
        "MHI_i": r.get("MHI_i"),
        "MHI_f": r.get("MHI_f"),
        "S_i": r.get("S_i"),
        "S_f": r.get("S_f"),
        "H": r.get("H"),
        "distance_km": r.get("distance_km"),
    })
dump("mhi_composite.json", composite)


# ------------------------------------------------------------
# 5. MHI_sf 前滩太古里候选 → mhi_sf.json
# ------------------------------------------------------------
print("\n[5/7] MHI_sf 候选 → mhi_sf.json")
wb = load("MHI_sf_前滩太古里候选商场空间功能相似度.xlsx")
ws_res = wb["MHI_sf结果"]
candidates = list(rows(ws_res))

ws_id = wb["楼层识别率"]
identify = list(rows(ws_id))

# 三家商场的楼层 × 业态矩阵
floor_matrices = {}
for sheet_name in ["矩阵_前滩太古里", "矩阵_环贸iapm商场", "矩阵_兴业太古汇"]:
    ws = wb[sheet_name]
    header = [c.value for c in ws[1]]
    categories = header[1:]  # 14 个业态
    levels = []
    matrix = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        levels.append(r[0])
        matrix.append(list(r[1:]))
    mall = sheet_name.replace("矩阵_", "")
    floor_matrices[mall] = {
        "levels": levels,
        "categories": categories,
        "matrix": matrix,
    }

dump("mhi_sf.json", {
    "target": "前滩太古里",
    "candidates": candidates,
    "identify_rate": identify,
})

dump("floor_matrix.json", floor_matrices)


# ------------------------------------------------------------
# 6. 商场_品牌.xlsx → brand_coverage.json（top 品牌覆盖榜）
# ------------------------------------------------------------
print("\n[6/7] 品牌覆盖 → brand_coverage.json")
wb = load("商场_品牌.xlsx")
ws = wb["品牌覆盖汇总"]
brand_cov = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[0] is None:
        continue
    brand_cov.append({
        "brand": r[0],
        "mall_count": r[1],
    })
# 仅保留 top 50（覆盖 ≥ 20 家的品牌）
brand_cov_top = sorted(brand_cov, key=lambda x: -(x["mall_count"] or 0))[:50]
dump("brand_coverage.json", brand_cov_top)


# ------------------------------------------------------------
# 7. 商场_功能.xlsx → function_summary.json（每个商场覆盖了哪些功能）
# ------------------------------------------------------------
print("\n[7/7] 商场功能清单 → function_summary.json")
wb = load("商场_功能.xlsx")
ws = wb["商场功能汇总"]
func_summary = []
for r in ws.iter_rows(min_row=3, values_only=True):  # 第 1 行是合并标题，第 2 行是表头
    if r[0] is None:
        continue
    func_summary.append({
        "name": r[0],
        "function_count": r[1],
        "functions": (r[2] or "").split("、") if r[2] else [],
    })
dump("function_summary.json", func_summary)



# ------------------------------------------------------------
# bonus: 生成 viz_quadrant 直接消费的接入数据 quadrant_phase1.json
# 把复合 MHI 表转成 viz_quadrant.js 期望的 {axes, malls} 形态
# ------------------------------------------------------------
print("\n[bonus] 生成 viz 接入数据 quadrant_phase1.json")
import statistics

ANCHOR_NAME = "前滩太古里"
HIGHEND_KEYWORDS = ("高端", "奢华", "luxury")

xs = [c["MHI_i"] for c in composite if c["MHI_i"] is not None]
ys = [c["MHI_f"] for c in composite if c["MHI_f"] is not None]
x_med, y_med = statistics.median(xs), statistics.median(ys)


def role_of(c):
    if c["name"] == ANCHOR_NAME:
        return "anchor"
    pos = c.get("positioning") or ""
    if any(k in pos for k in HIGHEND_KEYWORDS):
        return "highend"
    return "standard"


def short_id(name):
    """生成 viz 期望的紧凑 id"""
    return "".join(ch for ch in name if not ch.isspace())[:16]


quadrant_malls = []
for c in composite:
    if c["MHI_i"] is None or c["MHI_f"] is None:
        continue
    quadrant_malls.append({
        "id": short_id(c["name"]),
        "name": c["name"],
        "x": c["MHI_i"],
        "y": c["MHI_f"],
        "role": role_of(c),
        "positioning": c["positioning"],
        "circle": c["circle"],
        "district": c["district"],
        "brand_count": c["brand_count"],
        "function_count": c["function_count"],
        "distance_km": c["distance_km"],
        "H_rank": c["rank"],
        "note": f"{c['district']} · {c['circle']} · {c['positioning']} · 复合 MHI 排名 {c['rank']}",
    })


def padded_domain(values, pad=0.08):
    """给数据范围加 padding，避免点贴边"""
    lo, hi = min(values), max(values)
    span = hi - lo
    return [lo - span * pad, hi + span * pad]


# 6 档消费分层切片
SLICE_DEFS = [
    {"key": "total",     "label": "全部 80 家", "positionings": None},
    {"key": "mass",      "label": "大众 / 亲民", "positionings": ["大众/亲民"]},
    {"key": "mid",       "label": "中端",        "positionings": ["中端"]},
    {"key": "upper_mid", "label": "中高端",      "positionings": ["中高端"]},
    {"key": "high",      "label": "高端",        "positionings": ["高端"]},
    {"key": "outlet",    "label": "奥莱 / 折扣", "positionings": ["奥莱/折扣"]},
]
slices = []
for sd in SLICE_DEFS:
    if sd["positionings"] is None:
        cnt = len(quadrant_malls)
    else:
        cnt = sum(1 for m in quadrant_malls if m["positioning"] in sd["positionings"])
    slices.append({
        "key": sd["key"],
        "label": sd["label"],
        "positionings": sd["positionings"],
        "count": cnt,
    })

quadrant_data = {
    "axes": {
        "x_label": "MHI_i · 品牌池同质化指标",
        "y_label": "MHI_f · 业态向量同质化指标",
        "x_domain": padded_domain(xs),
        "y_domain": padded_domain(ys),
        "x_median": x_med,
        "y_median": y_med,
    },
    "malls": quadrant_malls,
    "slices": slices,
}
dump("quadrant_phase1.json", quadrant_data)
print(f"  axes: x_domain={[round(v,4) for v in quadrant_data['axes']['x_domain']]}, "
      f"y_domain={[round(v,3) for v in quadrant_data['axes']['y_domain']]}")
print(f"  锚点: 1，高端: {sum(1 for m in quadrant_malls if m['role']=='highend')}，"
      f"标准: {sum(1 for m in quadrant_malls if m['role']=='standard')}")


print(f"\n全部输出到 {OUT}")
print(f"文件清单：")
for f in sorted(OUT.glob("*.json")):
    print(f"  {f.name:40s} {f.stat().st_size/1024:>7.1f} KB")
